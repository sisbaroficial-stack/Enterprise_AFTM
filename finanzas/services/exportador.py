"""
SERVICIO DE EXPORTACIÓN
Exporta reportes financieros a Excel y PDF
"""

from io import BytesIO
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.platypus import Image as RLImage
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ExportadorExcel:
    """
    Exporta reportes financieros a Excel con formato profesional
    """
    
    @staticmethod
    def exportar_gastos(gastos, nombre_archivo='Gastos'):
        """
        Exporta lista de gastos a Excel con formato
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Gastos"
        
        # Título
        ws['A1'] = 'REPORTE DE GASTOS'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        
        # Fecha del reporte
        ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:H2')
        
        # Encabezados
        headers = ['Fecha', 'Categoría', 'Concepto', 'Sucursal', 'Monto', 'Estado', 'Registrado Por', 'Método Pago']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        row = 5
        total = Decimal('0')
        
        for gasto in gastos:
            ws.cell(row=row, column=1, value=gasto.fecha.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=2, value=gasto.categoria.nombre)
            ws.cell(row=row, column=3, value=gasto.concepto)
            ws.cell(row=row, column=4, value=gasto.sucursal.nombre if gasto.sucursal else 'General')
            ws.cell(row=row, column=5, value=float(gasto.monto))
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=6, value=gasto.get_estado_display())
            ws.cell(row=row, column=7, value=gasto.registrado_por.get_full_name())
            ws.cell(row=row, column=8, value=gasto.get_metodo_pago_display())
            
            total += gasto.monto
            row += 1
        
        # Total
        ws.cell(row=row, column=4, value='TOTAL:').font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(total))
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True, size=12)
        ws.cell(row=row, column=5).fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws[f'A4:H{row}']:
            for cell in row_cells:
                cell.border = thin_border
        
        # Crear respuesta HTTP
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    @staticmethod
    def exportar_reporte_financiero(datos, mes, anio, nombre_archivo='Reporte_Financiero'):
        """
        Exporta reporte financiero completo con gráficas
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl no está instalado")
        
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen Financiero"
        
        # Título
        ws1['A1'] = f'REPORTE FINANCIERO - {mes}/{anio}'
        ws1['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        ws1['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('A1:D1')
        
        # KPIs
        row = 3
        kpis = [
            ('💵 Ventas Totales:', float(datos['total_ventas'])),
            ('📦 Costo de Mercancía:', float(datos['costo_mercancia'])),
            ('✨ Utilidad Bruta:', float(datos['utilidad_bruta'])),
            ('💸 Gastos Operativos:', float(datos['total_gastos'])),
            ('✅ UTILIDAD NETA:', float(datos['utilidad_neta'])),
        ]
        
        for label, valor in kpis:
            ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws1.cell(row=row, column=2, value=valor).number_format = '$#,##0'
            
            if 'UTILIDAD NETA' in label:
                ws1.cell(row=row, column=2).font = Font(bold=True, size=14, color='008000')
                ws1.cell(row=row, column=2).fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            row += 1
        
        # Márgenes
        row += 1
        ws1.cell(row=row, column=1, value='Margen Bruto:').font = Font(bold=True)
        ws1.cell(row=row, column=2, value=float(datos['margen_bruto']) / 100).number_format = '0.00%'
        
        row += 1
        ws1.cell(row=row, column=1, value='Margen Neto:').font = Font(bold=True)
        ws1.cell(row=row, column=2, value=float(datos['margen_neto']) / 100).number_format = '0.00%'
        
        # Ajustar anchos
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        
        # Hoja 2: Gastos por Categoría
        ws2 = wb.create_sheet(title="Gastos por Categoría")
        
        ws2['A1'] = 'Categoría'
        ws2['B1'] = 'Total'
        ws2['C1'] = '% del Total'
        
        for col in ['A1', 'B1', 'C1']:
            ws2[col].font = Font(bold=True, color='FFFFFF')
            ws2[col].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        
        row = 2
        for item in datos.get('gastos_por_categoria', []):
            ws2.cell(row=row, column=1, value=item['categoria'].nombre)
            ws2.cell(row=row, column=2, value=float(item['total'])).number_format = '$#,##0'
            ws2.cell(row=row, column=3, value=float(item['porcentaje']) / 100).number_format = '0.00%'
            row += 1
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 12
        
        # Agregar gráfica de pastel
        if datos.get('gastos_por_categoria'):
            pie = PieChart()
            labels = Reference(ws2, min_col=1, min_row=2, max_row=row-1)
            data = Reference(ws2, min_col=2, min_row=1, max_row=row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución de Gastos"
            ws2.add_chart(pie, "E2")
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{mes}_{anio}.xlsx"'
        
        return response


class ExportadorPDF:
    """
    Exporta reportes financieros a PDF profesional
    """
    
    @staticmethod
    def exportar_reporte_financiero(datos, mes, anio, nombre_archivo='Reporte_Financiero'):
        """
        Genera PDF profesional del reporte financiero
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab no está instalado")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                              rightMargin=72, leftMargin=72,
                              topMargin=72, bottomMargin=18)
        
        elementos = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        titulo = Paragraph(f'REPORTE FINANCIERO<br/>{mes}/{anio}', titulo_style)
        elementos.append(titulo)
        elementos.append(Spacer(1, 12))
        
        # Fecha de generación
        fecha_generacion = Paragraph(
            f'<para align=center>Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}</para>',
            styles['Normal']
        )
        elementos.append(fecha_generacion)
        elementos.append(Spacer(1, 20))
        
        # Tabla de KPIs
        data_kpis = [
            ['CONCEPTO', 'VALOR'],
            ['💵 Ventas Totales', f"${float(datos['total_ventas']):,.0f}"],
            ['📦 Costo de Mercancía', f"-${float(datos['costo_mercancia']):,.0f}"],
            ['✨ Utilidad Bruta', f"${float(datos['utilidad_bruta']):,.0f}"],
            ['💸 Gastos Operativos', f"-${float(datos['total_gastos']):,.0f}"],
            ['', ''],
            ['✅ UTILIDAD NETA', f"${float(datos['utilidad_neta']):,.0f}"],
            ['Margen Neto', f"{float(datos['margen_neto']):.1f}%"],
        ]
        
        tabla_kpis = Table(data_kpis, colWidths=[3*inch, 2*inch])
        tabla_kpis.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 6), (-1, 6), colors.HexColor('#E2EFDA')),
            ('FONTNAME', (0, 6), (-1, 7), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 6), (-1, 6), 14),
            ('TEXTCOLOR', (0, 6), (-1, 6), colors.HexColor('#008000')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elementos.append(tabla_kpis)
        elementos.append(Spacer(1, 30))
        
        # Gastos por Categoría
        if datos.get('gastos_por_categoria'):
            titulo_cat = Paragraph('<para align=center><b>GASTOS POR CATEGORÍA</b></para>', styles['Heading2'])
            elementos.append(titulo_cat)
            elementos.append(Spacer(1, 12))
            
            data_cat = [['CATEGORÍA', 'MONTO', '% DEL TOTAL']]
            
            for item in datos['gastos_por_categoria']:
                data_cat.append([
                    item['categoria'].nombre,
                    f"${float(item['total']):,.0f}",
                    f"{float(item['porcentaje']):.1f}%"
                ])
            
            tabla_cat = Table(data_cat, colWidths=[3*inch, 1.5*inch, 1.5*inch])
            tabla_cat.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ]))
            
            elementos.append(tabla_cat)
        
        # Construir PDF
        doc.build(elementos)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}_{mes}_{anio}.pdf"'
        
        return response
    
    @staticmethod
    def exportar_desprendible_nomina(nomina):
        """
        Genera desprendible de nómina en PDF
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab no está instalado")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo = Paragraph(
            '<para align=center><b>DESPRENDIBLE DE PAGO</b></para>',
            styles['Heading1']
        )
        elementos.append(titulo)
        elementos.append(Spacer(1, 20))
        
        # Información del empleado
        info_empleado = [
            ['Empleado:', nomina.empleado.nombre_completo],
            ['Documento:', nomina.empleado.numero_documento],
            ['Cargo:', nomina.empleado.cargo],
            ['Período:', f'{nomina.mes}/{nomina.anio}'],
        ]
        
        tabla_info = Table(info_empleado, colWidths=[1.5*inch, 4*inch])
        tabla_info.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        elementos.append(tabla_info)
        elementos.append(Spacer(1, 20))
        
        # Devengado
        data_devengado = [
            ['DEVENGADO', ''],
            ['Salario Base', f'${float(nomina.salario_base):,.0f}'],
            ['Auxilio Transporte', f'${float(nomina.auxilio_transporte):,.0f}'],
            ['Horas Extras', f'${float(nomina.horas_extras):,.0f}'],
            ['Bonificaciones', f'${float(nomina.bonificaciones):,.0f}'],
            ['TOTAL DEVENGADO', f'${float(nomina.total_devengado):,.0f}'],
        ]
        
        tabla_devengado = Table(data_devengado, colWidths=[3*inch, 2*inch])
        tabla_devengado.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#E2EFDA')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elementos.append(tabla_devengado)
        elementos.append(Spacer(1, 15))
        
        # Deducciones
        data_deducciones = [
            ['DEDUCCIONES', ''],
            ['Salud (4%)', f'${float(nomina.salud):,.0f}'],
            ['Pensión (4%)', f'${float(nomina.pension):,.0f}'],
            ['Préstamos', f'${float(nomina.prestamos):,.0f}'],
            ['Otras Deducciones', f'${float(nomina.otras_deducciones):,.0f}'],
            ['TOTAL DEDUCCIONES', f'${float(nomina.total_deducciones):,.0f}'],
        ]
        
        tabla_deducciones = Table(data_deducciones, colWidths=[3*inch, 2*inch])
        tabla_deducciones.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C00000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#FFE7E6')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        
        elementos.append(tabla_deducciones)
        elementos.append(Spacer(1, 20))
        
        # Neto a pagar
        data_neto = [
            ['NETO A PAGAR', f'${float(nomina.neto_pagar):,.0f}'],
        ]
        
        tabla_neto = Table(data_neto, colWidths=[3*inch, 2*inch])
        tabla_neto.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        
        elementos.append(tabla_neto)
        
        # Construir
        doc.build(elementos)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Desprendible_{nomina.empleado.numero_documento}_{nomina.mes}_{nomina.anio}.pdf"'
        
        return response