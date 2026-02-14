from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Q
from decimal import Decimal
import io
from django.utils import timezone
from django.db.models import Avg, Sum, F
from datetime import timedelta
from .models import SugerenciaCompra, ConfiguracionCompras
from .services import ServicioSugerenciasCompra
from inventario.models import Producto, InventarioSucursal, MovimientoInventario

from usuarios.views import registrar_actividad

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# PDF
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


def obtener_sucursal_usuario(request):
    """Helper para obtener sucursal del usuario"""
    from inventario.views import obtener_sucursal_usuario as get_sucursal
    return get_sucursal(request)


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

@login_required
def sugerencias_compra(request):
    """Vista principal de sugerencias de compra"""
    
    # Solo SUPER_ADMIN y ADMIN
    if request.user.rol not in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, '⛔ No tienes permisos para acceder a sugerencias de compra')
        return redirect('dashboard:home')
    
    sucursal = obtener_sucursal_usuario(request)
    from django.http import HttpResponseRedirect
    if isinstance(sucursal, HttpResponseRedirect):
        return sucursal
    
    # Generar sugerencias si se solicita
    if request.GET.get('generar') == 'true':
        try:
            servicio = ServicioSugerenciasCompra(sucursal=sucursal, usuario=request.user)
            sugerencias_creadas = servicio.generar_sugerencias_todas()
            
            registrar_actividad(
                request.user,
                tipo='CREAR',
                descripcion=f'Generó {len(sugerencias_creadas)} sugerencias de compra con IA',
                request=request
            )
            
            messages.success(
                request, 
                f'✅ Se generaron {len(sugerencias_creadas)} sugerencias usando Inteligencia Artificial'
            )
        except Exception as e:
            messages.error(request, f'❌ Error al generar sugerencias: {str(e)}')
        
        return redirect('compras:sugerencias_compra')
    
    # Obtener sugerencias existentes
    if sucursal:
        sugerencias = SugerenciaCompra.objects.filter(sucursal=sucursal)
    else:
        sugerencias = SugerenciaCompra.objects.all()
    
    # Filtros
    urgencia_filtro = request.GET.get('urgencia', '')
    proveedor_filtro = request.GET.get('proveedor', '')
    busqueda = request.GET.get('q', '')
    
    if urgencia_filtro:
        sugerencias = sugerencias.filter(urgencia=urgencia_filtro)
    
    if proveedor_filtro:
        sugerencias = sugerencias.filter(proveedor_sugerido_id=proveedor_filtro)
    
    if busqueda:
        sugerencias = sugerencias.filter(
            Q(producto__nombre__icontains=busqueda) |
            Q(producto__codigo__icontains=busqueda)
        )
    
    sugerencias = sugerencias.select_related(
        'producto', 
        'producto__categoria', 
        'proveedor_sugerido',
        'sucursal'
    ).order_by('urgencia', 'dias_stock_restante')
    
    # Estadísticas (calculadas ANTES de paginar)
    stats = {
        'total': sugerencias.count(),
        'urgentes': sugerencias.filter(urgencia='URGENTE').count(),
        'altas': sugerencias.filter(urgencia='ALTA').count(),
        'medias': sugerencias.filter(urgencia='MEDIA').count(),
        'bajas': sugerencias.filter(urgencia='BAJA').count(),
        'inversion_total': sugerencias.aggregate(
            total=Sum('inversion_estimada')
        )['total'] or Decimal('0'),
        'confianza_promedio': round(
            sugerencias.aggregate(Avg('confianza_ia'))['confianza_ia__avg'] or 0, 
            1
        ),
        'productos_crecientes': sugerencias.filter(tendencia='CRECIENTE').count(),
        'productos_estables': sugerencias.filter(tendencia='ESTABLE').count(),
        'productos_decrecientes': sugerencias.filter(tendencia='DECRECIENTE').count(),
    }
    
    # Agrupar por urgencia (para las tabs)
    urgentes_qs = sugerencias.filter(urgencia='URGENTE')
    altas_qs = sugerencias.filter(urgencia='ALTA')
    
    # PAGINACIÓN PARA CADA TAB
    # Tab "Urgentes"
    page_urgentes = request.GET.get('page_urgentes', 1)
    paginator_urgentes = Paginator(urgentes_qs, 20)  # 20 por página
    try:
        urgentes = paginator_urgentes.page(page_urgentes)
    except PageNotAnInteger:
        urgentes = paginator_urgentes.page(1)
    except EmptyPage:
        urgentes = paginator_urgentes.page(paginator_urgentes.num_pages)
    
    # Tab "Altas"
    page_altas = request.GET.get('page_altas', 1)
    paginator_altas = Paginator(altas_qs, 20)
    try:
        altas = paginator_altas.page(page_altas)
    except PageNotAnInteger:
        altas = paginator_altas.page(1)
    except EmptyPage:
        altas = paginator_altas.page(paginator_altas.num_pages)
    
    # Tab "Todas" (tabla completa)
    page = request.GET.get('page', 1)
    paginator_todas = Paginator(sugerencias, 50)  # 50 por página
    try:
        sugerencias_paginadas = paginator_todas.page(page)
    except PageNotAnInteger:
        sugerencias_paginadas = paginator_todas.page(1)
    except EmptyPage:
        sugerencias_paginadas = paginator_todas.page(paginator_todas.num_pages)
    
    # Proveedores únicos para filtro
    from proveedores.models import Proveedor
    proveedores = Proveedor.objects.filter(
        activo=True,
        productos__sugerencias_compra__isnull=False
    ).distinct()
    
    # Configuración
    config = ConfiguracionCompras.objects.first()
    
    context = {
            'sugerencias': sugerencias_paginadas,  # ← Paginado para tabla
            'todas_sugerencias': list(sugerencias[:100]),  # ← Para gráficas (primeras 100)
            'urgentes': urgentes,  # ← Paginado
            'altas': altas,  # ← Paginado
            'stats': stats,
            'proveedores': proveedores,
            'config': config,
            'sucursal': sucursal,
    }
    
    return render(request, 'compras/sugerencias.html', context)


@login_required
def detalle_sugerencia(request, sugerencia_id):
    """Ver detalle completo de una sugerencia con histórico de ventas"""
    
    if request.user.rol not in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, '⛔ No tienes permisos')
        return redirect('dashboard:home')
    
    # Obtener la sugerencia
    sugerencia = get_object_or_404(SugerenciaCompra, pk=sugerencia_id)

    # Histórico de ventas (últimos 90 días)
    from django.utils import timezone
    fecha_inicio = timezone.now().date() - timedelta(days=90)

    historico = MovimientoInventario.objects.filter(
        producto=sugerencia.producto,
        tipo='SALIDA',
        motivo='VENTA',
        fecha__gte=fecha_inicio
    )

    if sugerencia.sucursal:
        historico = historico.filter(sucursal=sugerencia.sucursal)

    historico = historico.values('fecha').annotate(
        dia=F('fecha'),
        total=Sum('cantidad')
    ).order_by('-fecha')

    context = {
        'sugerencia': sugerencia,
        'historico': historico,
    }

    return render(request, 'compras/detalle_sugerencia.html', context)

@login_required
def exportar_excel(request):
    """Exportar sugerencias a Excel profesional"""
    
    if request.user.rol not in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, '⛔ No tienes permisos')
        return redirect('dashboard:home')
    
    sucursal = obtener_sucursal_usuario(request)
    
    # Obtener sugerencias
    if sucursal:
        sugerencias = SugerenciaCompra.objects.filter(sucursal=sucursal)
    else:
        sugerencias = SugerenciaCompra.objects.all()
    
    sugerencias = sugerencias.select_related(
        'producto',
        'proveedor_sugerido',
        'sucursal'
    ).order_by('urgencia', 'dias_stock_restante')
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sugerencias de Compra"
    
    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    urgente_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    alta_fill = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
    media_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    baja_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'Urgencia', 'Producto', 'Código', 'Proveedor', 'Stock Actual',
        'Promedio Ventas/Día', 'Días Stock', 'Predicción 30d (IA)',
        'Tendencia', 'Cantidad Sugerida', 'Costo Unit.', 'Inversión Total',
        'Punto Reorden', 'Confianza IA %', 'Razón'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Datos
    for row_num, sug in enumerate(sugerencias, 2):
        # Determinar color de fila
        if sug.urgencia == 'URGENTE':
            fill = urgente_fill
        elif sug.urgencia == 'ALTA':
            fill = alta_fill
        elif sug.urgencia == 'MEDIA':
            fill = media_fill
        else:
            fill = baja_fill
        
        row_data = [
            sug.get_urgencia_display(),
            sug.producto.nombre,
            sug.producto.codigo,
            sug.proveedor_sugerido.nombre if sug.proveedor_sugerido else 'Sin proveedor',
            sug.stock_actual,
            float(sug.promedio_ventas_diarias),
            float(sug.dias_stock_restante),
            sug.prediccion_proximos_30_dias,
            sug.get_tendencia_display(),
            sug.cantidad_sugerida,
            float(sug.costo_unitario),
            float(sug.inversion_estimada),
            sug.punto_reorden,
            float(sug.confianza_ia),
            sug.razon
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar anchos
    column_widths = [15, 30, 12, 20, 12, 18, 12, 18, 15, 18, 12, 15, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Guardar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Registrar actividad
    registrar_actividad(
        request.user,
        tipo='EXPORTAR',
        descripcion=f'Exportó {sugerencias.count()} sugerencias de compra a Excel',
        request=request
    )
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=sugerencias_compra_{sucursal.codigo if sucursal else "global"}.xlsx'
    
    return response


@login_required
def exportar_pdf(request):
    """Exportar sugerencias a PDF profesional"""
    
    if request.user.rol not in ['SUPER_ADMIN', 'ADMIN']:
        messages.error(request, '⛔ No tienes permisos')
        return redirect('dashboard:home')
    
    sucursal = obtener_sucursal_usuario(request)
    
    # Obtener sugerencias urgentes y altas
    if sucursal:
        sugerencias = SugerenciaCompra.objects.filter(
            sucursal=sucursal,
            urgencia__in=['URGENTE', 'ALTA']
        )
    else:
        sugerencias = SugerenciaCompra.objects.filter(urgencia__in=['URGENTE', 'ALTA'])
    
    sugerencias = sugerencias.select_related(
        'producto',
        'proveedor_sugerido'
    ).order_by('urgencia', 'dias_stock_restante')[:50]
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#667EEA'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    elements.append(Paragraph("📊 SUGERENCIAS DE COMPRA", title_style))
    elements.append(Paragraph(
        f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    ))
    
    if sucursal:
        elements.append(Paragraph(f"Sucursal: {sucursal.nombre}", styles['Normal']))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabla
    data = [['Producto', 'Stock', 'Días', 'Sugerido', 'Inversión', 'Urgencia']]
    
    for sug in sugerencias:
        data.append([
            sug.producto.nombre[:30],
            str(sug.stock_actual),
            f"{float(sug.dias_stock_restante):.1f}",
            str(sug.cantidad_sugerida),
            f"${float(sug.inversion_estimada):,.0f}",
            sug.get_urgencia_display()[:10]
        ])
    
    table = Table(data, colWidths=[3*inch, 0.8*inch, 0.8*inch, 1*inch, 1.2*inch, 1.2*inch])
    
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Resumen
    elements.append(Spacer(1, 0.3*inch))
    total_inversion = sum(float(s.inversion_estimada) for s in sugerencias)
    elements.append(Paragraph(
        f"<b>INVERSIÓN TOTAL ESTIMADA: ${total_inversion:,.0f} COP</b>",
        styles['Heading2']
    ))
    
    # Construir PDF
    doc.build(elements)
    
    # Registrar actividad
    registrar_actividad(
        request.user,
        tipo='EXPORTAR',
        descripcion=f'Exportó {sugerencias.count()} sugerencias a PDF',
        request=request
    )
    
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=sugerencias_{sucursal.codigo if sucursal else "global"}.pdf'
    
    return response


@login_required
def configuracion_compras(request):
    """Configurar parámetros del módulo de compras"""
    
    if request.user.rol != 'SUPER_ADMIN':
        messages.error(request, '⛔ Solo SUPER_ADMIN puede cambiar configuración')
        return redirect('compras:sugerencias_compra')
    
    config, created = ConfiguracionCompras.objects.get_or_create(pk=1)
    
    if request.method == 'POST':
        try:
            config.dias_cobertura_default = int(request.POST.get('dias_cobertura', 30))
            config.stock_seguridad_porcentaje = int(request.POST.get('stock_seguridad', 20))
            config.dias_analisis_historico = int(request.POST.get('dias_analisis', 90))
            config.umbral_urgente_dias = int(request.POST.get('umbral_urgente', 3))
            config.umbral_alta_dias = int(request.POST.get('umbral_alta', 7))
            config.umbral_media_dias = int(request.POST.get('umbral_media', 15))
            config.habilitar_ia = request.POST.get('habilitar_ia') == 'on'
            config.save()
            
            messages.success(request, '✅ Configuración actualizada correctamente')
            return redirect('compras:sugerencias_compra')
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
    
    return render(request, 'compras/configuracion.html', {'config': config})