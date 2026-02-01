from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Sum, Count, Q, F, Avg
from decimal import Decimal

from inventario.models import Producto, InventarioSucursal, MovimientoInventario
from categorias.models import Categoria

from usuarios.views import registrar_actividad


@login_required
def reportes_home_view(request):
    """Página principal de reportes"""
    total_productos = Producto.objects.filter(activo=True).count()
    categorias_count = Categoria.objects.filter(activa=True).count()
    hace_30_dias = timezone.now() - timedelta(days=30)
    movimientos_mes = MovimientoInventario.objects.filter(fecha__gte=hace_30_dias).count()

    context = {
        'total_productos': total_productos,
        'categorias_count': categorias_count,
        'movimientos_mes': movimientos_mes,
    }
    return render(request, 'reportes/home.html', context)


@login_required
def exportar_productos_excel(request):
    """Exportar inventario completo a Excel CON STOCK REAL"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws.merge_cells('A1:J1')
    ws['A1'].value = "REPORTE DE INVENTARIO - SISBAR"
    ws['A1'].font = Font(bold=True, size=16, color="667EEA")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'].value = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Encabezados
    ws.append([])
    headers = ['Código', 'Nombre', 'Categoría', 'Subcategoría', 'Stock Total', 
               'Unidad', 'Min', 'Estado', 'Precio Compra', 'Proveedor']
    ws.append(headers)

    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # ✅ CALCULAR STOCK REAL DE TODAS LAS SUCURSALES
    productos = Producto.objects.filter(activo=True).select_related('categoria', 'subcategoria', 'proveedor')
    
    for p in productos:
        # Sumar stock de todas las sucursales
        stock_total = InventarioSucursal.objects.filter(
            producto=p
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        cantidad_minima_total = InventarioSucursal.objects.filter(
            producto=p
        ).aggregate(total=Sum('cantidad_minima'))['total'] or 0
        
        # Determinar estado
        if stock_total == 0:
            estado = '🔴 Agotado'
        elif stock_total <= cantidad_minima_total:
            estado = '🟡 Por Agotarse'
        else:
            estado = '🟢 Disponible'
        
        ws.append([
            p.codigo,
            p.nombre,
            p.categoria.nombre,
            p.subcategoria.nombre if p.subcategoria else 'N/A',
            stock_total,  # ✅ STOCK REAL
            p.get_unidad_medida_display(),
            cantidad_minima_total,
            estado,
            float(p.precio_compra),
            p.proveedor.nombre if p.proveedor else 'N/A'
        ])

    # Bordes y formatos
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = border
            if cell.column in [5, 7]:  # Cantidad y Mínimo
                cell.alignment = Alignment(horizontal='center')
            if cell.column == 9:  # Precio
                cell.number_format = '$#,##0.00'

    # Ajustar anchos
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 25

    # Registrar actividad
    registrar_actividad(request.user, 'EXPORTAR', 'Exportó inventario a Excel', request)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_movimientos_excel(request):
    """Exportar movimientos a Excel según rango de días"""
    dias = int(request.GET.get('dias', 30))
    fecha_desde = timezone.now() - timedelta(days=dias)

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells('A1:H1')
    ws['A1'].value = f"REPORTE DE MOVIMIENTOS - Últimos {dias} días"
    ws['A1'].font = Font(bold=True, size=16, color="667EEA")
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append([])
    headers = ['Fecha', 'Tipo', 'Producto', 'Código', 'Cantidad', 'Motivo', 'Usuario', 'Sucursal']
    ws.append(headers)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    movimientos = MovimientoInventario.objects.filter(
        fecha__gte=fecha_desde
    ).select_related('producto', 'usuario', 'sucursal').order_by('-fecha')

    for m in movimientos:
        ws.append([
            m.fecha.strftime('%d/%m/%Y %H:%M'),
            m.get_tipo_display(),
            m.producto.nombre,
            m.producto.codigo,
            m.cantidad,
            m.get_motivo_display() if hasattr(m, 'get_motivo_display') else m.motivo,
            m.usuario.username if m.usuario else "Sistema",
            m.sucursal.nombre
        ])

    # Bordes
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border

    # Ajustar anchos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20

    registrar_actividad(request.user, 'EXPORTAR', f'Exportó movimientos ({dias} días) a Excel', request)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


# ===============================
# REPORTE ABC
# ===============================
@login_required
def reporte_abc(request):
    """Reporte ABC - Clasificación de productos por ventas"""
    dias = int(request.GET.get('dias', 30))
    fecha_desde = timezone.now() - timedelta(days=dias)
    
    # Obtener ventas por producto
    ventas_por_producto = MovimientoInventario.objects.filter(
        fecha__gte=fecha_desde,
        tipo='SALIDA',
        motivo='VENTA'
    ).values('producto').annotate(
        total_vendido=Sum('cantidad'),
        valor_total=Sum(F('cantidad') * F('producto__precio_compra'))
    ).order_by('-valor_total')
    
    # Calcular total de ventas
    total_ventas = sum(item['valor_total'] or 0 for item in ventas_por_producto)
    
    # Clasificar productos
    productos_clasificados = []
    acumulado = Decimal('0')
    
    for item in ventas_por_producto:
        producto = Producto.objects.get(id=item['producto'])
        valor = item['valor_total'] or Decimal('0')
        porcentaje = (valor / total_ventas * 100) if total_ventas > 0 else 0
        acumulado += porcentaje
        
        # Clasificación ABC
        if acumulado <= 80:
            clase = 'A'
            color = 'success'
        elif acumulado <= 95:
            clase = 'B'
            color = 'warning'
        else:
            clase = 'C'
            color = 'secondary'
        
        productos_clasificados.append({
            'producto': producto,
            'total_vendido': item['total_vendido'],
            'valor_total': valor,
            'porcentaje': round(porcentaje, 2),
            'acumulado': round(acumulado, 2),
            'clase': clase,
            'color': color,
        })
    
    # Estadísticas por clase
    clase_a = sum(1 for p in productos_clasificados if p['clase'] == 'A')
    clase_b = sum(1 for p in productos_clasificados if p['clase'] == 'B')
    clase_c = sum(1 for p in productos_clasificados if p['clase'] == 'C')
    
    context = {
        'productos': productos_clasificados,
        'dias': dias,
        'total_ventas': total_ventas,
        'clase_a': clase_a,
        'clase_b': clase_b,
        'clase_c': clase_c,
    }
    return render(request, 'reportes/reporte_abc.html', context)


# ===============================
# REPORTE DE ROTACIÓN
# ===============================
@login_required
def reporte_rotacion(request):
    """Reporte de rotación de inventario"""
    dias = int(request.GET.get('dias', 30))
    fecha_desde = timezone.now() - timedelta(days=dias)
    
    productos_rotacion = []
    
    for producto in Producto.objects.filter(activo=True):
        # Ventas del período
        ventas = MovimientoInventario.objects.filter(
            producto=producto,
            tipo='SALIDA',
            motivo='VENTA',
            fecha__gte=fecha_desde
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        # Stock promedio (simplificado: stock actual)
        stock_actual = InventarioSucursal.objects.filter(
            producto=producto
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        if stock_actual > 0:
            rotacion = round(ventas / stock_actual, 2)
        else:
            rotacion = 0 if ventas == 0 else 999  # Sin stock pero con ventas = problema
        
        # Clasificación
        if rotacion >= 6:
            nivel = 'ALTA'
            color = 'success'
            recomendacion = 'Excelente - Mantener siempre en stock'
        elif rotacion >= 2:
            nivel = 'MEDIA'
            color = 'primary'
            recomendacion = 'Bueno - Control normal'
        elif rotacion >= 0.5:
            nivel = 'BAJA'
            color = 'warning'
            recomendacion = 'Regular - Considerar reducir stock'
        else:
            nivel = 'MUY BAJA'
            color = 'danger'
            recomendacion = 'Crítico - Liquidar o dejar de comprar'
        
        if ventas > 0 or stock_actual > 0:  # Solo mostrar productos con actividad
            productos_rotacion.append({
                'producto': producto,
                'ventas': ventas,
                'stock_actual': stock_actual,
                'rotacion': rotacion,
                'nivel': nivel,
                'color': color,
                'recomendacion': recomendacion,
            })
    
    # Ordenar por rotación descendente
    productos_rotacion.sort(key=lambda x: x['rotacion'], reverse=True)
    
    context = {
        'productos': productos_rotacion,
        'dias': dias,
    }
    return render(request, 'reportes/reporte_rotacion.html', context)


# ===============================
# PRODUCTOS SIN MOVIMIENTO
# ===============================
@login_required
def reporte_sin_movimiento(request):
    """Productos sin movimiento en X días"""
    dias = int(request.GET.get('dias', 30))
    fecha_desde = timezone.now() - timedelta(days=dias)
    
    # Productos que tienen movimientos en el período
    productos_con_movimiento = MovimientoInventario.objects.filter(
        fecha__gte=fecha_desde
    ).values_list('producto_id', flat=True).distinct()
    
    # Productos sin movimiento
    productos_sin_movimiento = Producto.objects.filter(
        activo=True
    ).exclude(
        id__in=productos_con_movimiento
    ).select_related('categoria', 'proveedor')
    
    # Calcular stock total y valor
    productos_info = []
    valor_total_inmovilizado = Decimal('0')
    
    for producto in productos_sin_movimiento:
        stock_total = InventarioSucursal.objects.filter(
            producto=producto
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        valor_stock = stock_total * producto.precio_compra
        valor_total_inmovilizado += valor_stock
        
        productos_info.append({
            'producto': producto,
            'stock_total': stock_total,
            'valor_stock': valor_stock,
        })
    
    # Ordenar por valor descendente
    productos_info.sort(key=lambda x: x['valor_stock'], reverse=True)
    
    context = {
        'productos': productos_info,
        'dias': dias,
        'total_productos': len(productos_info),
        'valor_total': valor_total_inmovilizado,
    }
    return render(request, 'reportes/sin_movimiento.html', context)